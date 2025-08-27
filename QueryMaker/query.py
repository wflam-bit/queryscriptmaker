import openpyxl

# Weekdays like "MON_FRI", "SAT", "SUN"
weekdays = "MON_FIR"
# Define the player_id 
player_id = "player_ID"
f= open("insert_query_{}_{}.sql".format(player_id, weekdays), "w+")
# Query, don't remove the {}
query = "INSERT INTO `metro-arrival`.arrival_time (arrival_day, time, player_id) VALUES ('{}', '{}', '{}');\n"
# Time Schedule Excel file must be located in the same folder as this script
dataframe = openpyxl.load_workbook("file.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        f.write(query.format(weekdays, col[row].value, player_id))

f.close()