import openpyxl
import customtkinter
from tkinter import filedialog as fd
import os
import subprocess
customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")
root = customtkinter.CTk()
root.geometry("500x500")

selected_filename = ""  # Variable to store the selected filename

# File selection, all type files / Could change to only ./xls files
def select_file():
    filetypes = (
        ('text files', '*.*'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)
    global selected_filename  # Use the global variable
    selected_filename = filename

# Create the queries from the schedule data
def createQuery():
    current_dir = os.getcwd()
    
    print("here "+current_dir)
    filename = selected_filename
    player_id = entry_1.get()
    weekdays = radiobutton_var.get()

    f = open("insert_query_{}_{}.sql".format(player_id, weekdays), "w+")
    query = "INSERT INTO `metro-arrival`.arrival_time (arrival_day, time, player_id) VALUES ('{}', '{}', '{}');\n"
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active

    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            f.write(query.format(weekdays, col[row].value, player_id))

   
    f.close()
    # open the relative path where this files is located
    path = os.path.realpath(current_dir)
    os.startfile(path)  
    

# CTk GUI
frame = customtkinter.CTkFrame(master=root)
frame.pack(pady=20, padx=60, fill="both", expand=True)

label = customtkinter.CTkLabel(master=frame, text="Query maker V3", font=(None))
label.pack(pady=12, padx=10)

entry_1 = customtkinter.CTkEntry(master=frame, placeholder_text="Player ID")
entry_1.pack(pady=10, padx=10)

radiobutton_var = customtkinter.StringVar(value="MON_FRI")

radiobutton_1 = customtkinter.CTkRadioButton(master=frame, text="MONDAY - FRIDAY", variable=radiobutton_var, value="MON_FRI")
radiobutton_1.pack(pady=10, padx=10)

radiobutton_2 = customtkinter.CTkRadioButton(master=frame, text="ONLY - SATURDAY", variable=radiobutton_var, value="SAT")
radiobutton_2.pack(pady=10, padx=8)

radiobutton_3 = customtkinter.CTkRadioButton(master=frame, text="ONLY - SUNDAY", variable=radiobutton_var, value="SUN")
radiobutton_3.pack(pady=10, padx=8)

button = customtkinter.CTkButton(master=frame, text="Open schedule file", command=select_file)
button.pack(pady=12, padx=10)

button2 = customtkinter.CTkButton(master=frame, text="Create", command=createQuery)
button2.pack(pady=12, padx=10)

root.mainloop()